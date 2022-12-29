import csv
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib
import numpy as np

designation = 0
salaryfrom = 1
salaryto = 2
sal = 3
areaname = 4
pubtime = 5

money = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

class Report:
    def __init__(self, filename, name):
        self.filename = filename
        self.name = name
        self.years = list(range(2007, 2023))
        self.sums = {}
        self.length = {}
        self.sumsCur = {}
        self.lengthCur = {}
        self.cities = []
        self.citySums = {}
        self.cities_length = {}
        self.vacancies_length = 0
        self.ansSums = {}
        self.part = {}
        self.read_file()
        self.calculate_file()
        self.workBook = Workbook()

    def read_file(self):
        first = False
        with open(self.filename, encoding="utf-8") as file:
            reader = csv.reader(file)
            for row in reader:
                if not first:
                    first = True
                    NAME = row.index("name")
                    SALARY_FROM = row.index("salary_from")
                    SALARY_TO = row.index("salary_to")
                    SALARY_CURRENCY = row.index("salary_currency")
                    AREA_NAME = row.index("area_name")
                    PUBLISHED_AT = row.index("published_at")
                else:
                    my_row = row.copy()
                    if all(my_row):
                        cur_year = int(row[PUBLISHED_AT].split("-")[0])
                        cur_salary = (int(float(row[SALARY_TO])) + int(float(row[SALARY_FROM]))) * money[
                            row[SALARY_CURRENCY]] // 2
                        cur_name = row[NAME]
                        cur_city = row[AREA_NAME]
                        self.sums[cur_year] = self.sums.get(cur_year, 0) + cur_salary
                        self.length[cur_year] = self.length.get(cur_year, 0) + 1
                        if profession in cur_name:
                            self.sumsCur[cur_year] = self.sumsCur.get(cur_year, 0) + cur_salary
                            self.lengthCur[cur_year] = self.lengthCur.get(cur_year, 0) + 1
                        if cur_city not in self.cities:
                            self.cities.append(cur_city)
                        self.citySums[cur_city] = self.citySums.get(cur_city, 0) + cur_salary
                        self.cities_length[cur_city] = self.cities_length.get(cur_city, 0) + 1
                        self.vacancies_length += 1

    def calculate_file(self):
        for i in self.years:
            if self.sums.get(i, None):
                self.sums[i] = int(self.sums[i] // self.length[i])
            if self.sumsCur.get(i, None):
                self.sumsCur[i] = int(self.sumsCur[i] // self.lengthCur[i])

        for i in self.cities:
            self.citySums[i] = int(self.citySums[i] // self.cities_length[i])
        neededCities = [city for city in self.cities if self.cities_length[city] >= self.vacancies_length // 100]
        self.ansSums = {key: self.citySums[key] for key in sorted(neededCities, key=lambda x: self.citySums[x], reverse=True)[:10]}
        self.part = {key: float("{:.4f}".format(self.cities_length[key] / self.vacancies_length)) for key in sorted(neededCities, key=lambda x: self.cities_length[x] / self.vacancies_length, reverse=True)[:10]}

    def print_file(self):
        print("Динамика уровня зарплат по годам:", self.sums)
        print("Динамика количества вакансий по годам:", self.length)
        if not len(self.sumsCur):
            self.sumsCur[2022] = 0
        print("Динамика уровня зарплат по годам для выбранной профессии:", self.sumsCur)
        if not len(self.lengthCur):
            self.lengthCur[2022] = 0
        print("Динамика количества вакансий по годам для выбранной профессии:", self.lengthCur)
        print("Уровень зарплат по городам (в порядке убывания):", self.ansSums)
        print("Доля вакансий по городам (в порядке убывания):", self.part)

    def generate_excel(self):
        self.yearsStatSheet = self.workBook.create_sheet(title="Статистика по годам")
        self.citiesStatSheet = self.workBook.create_sheet(title="Статистика по городам")
        self.workBook.remove(self.workBook["Sheet"])
        side = Side(border_style='thin', color="000000")
        self.border = Border(right=side, top=side, bottom=side, left=side)
        self.alignment = Alignment(horizontal='left')
        self.dataAlignment = Alignment(horizontal='right')
        self.citiesStatSheet["a1"] = 12
        self.report_years()
        self.report_cities()
        self.fit_cells()
        self.workBook.save('report.xlsx')

    def report_years(self):
        titles = ["Год", "Средняя зарплата", "Средняя зарплата - " + self.name,
                   "Количество вакансий", "Количество вакансий - " + self.name]
        self.set_headers(self.yearsStatSheet, titles)

        mtrx = []
        for values in range(len(self.sums)):
            data = list(self.sums.keys())[values]
            append = [data, self.sums[data], self.sumsCur[data], self.length[data],
                      self.lengthCur[data]]
            mtrx.append(append)

        self.fill_matrix(self.yearsStatSheet, mtrx, offset=(0, 1))

    def fill_matrix(self, sheet, matrix, offset=(0, 0)):
        for i in range(len(matrix)):
            for j in range(len(matrix[0])):
                address = f"{get_column_letter(j + 1 + offset[0])}{i + 1 + offset[1]}"
                sheet[address] = matrix[i][j]
                sheet[address].border = self.border
                sheet[address].alignment = self.dataAlignment
                sheet.column_dimensions[get_column_letter(j + 1)].auto_size = 1

    def set_headers(self, sheet, headers, offset=(0, 0)):
        for i in range(0, len(headers)):
            address = f"{get_column_letter(i + 1 + offset[0])}{1 + offset[1]}"
            sheet[address] = headers[i]
            sheet[address].border = self.border
            sheet[address].alignment = self.alignment
            sheet[address].font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(i + 1)].auto_size = 1

    def fit_cells(self):
        for sheetName in self.workBook.sheetnames:
            sheetWB = self.workBook[sheetName]
            for i in range(1, sheetWB.max_column + 1):
                width = None
                for j in range(1, sheetWB.max_row + 1):
                    value = sheetWB[f"{get_column_letter(i)}{j}"].value
                    if value is not None and (width is None or len(str(value)) > width):
                        width = len(str(value))
                if width is not None:
                    sheetWB.column_dimensions[f"{get_column_letter(i)}"].width = width + 2
                else:
                    sheetWB.column_dimensions[f"{get_column_letter(i)}"].width = + 2

    def report_cities(self):
        salaryLevel = ["Город", "Уровень зарплат"]
        shareOfVacancies = ["Город", "Доля вакансий"]
        self.set_headers(self.citiesStatSheet, salaryLevel)
        self.set_headers(self.citiesStatSheet, shareOfVacancies, (3, 0))

        self.dataAlignment = Alignment(horizontal='left')
        self.fill_matrix(self.citiesStatSheet, [[i] for i in self.ansSums.keys()], offset=(0, 1))
        max = {key: f"{(val * 10000) // 1 / 100}%" for key, val in self.part.items()}
        self.fill_matrix(self.citiesStatSheet, [[i] for i in list(max.keys())], offset=(3, 1))
        self.dataAlignment = Alignment(horizontal='right')
        self.fill_matrix(self.citiesStatSheet, [[i] for i in list(self.ansSums.values())], offset=(1, 1))
        self.fill_matrix(self.citiesStatSheet, [[i] for i in list(max.values())], offset=(4, 1))

    def generate_image(self):
        matplotlib.rc("font", size=8)
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)
        width = 0.3
        x = np.arange(len(self.sums.keys()))
        remuneration1 = ax1.bar(x - width / 2, self.sums.values(), width, label="средняя з/п")
        remuneration2 = ax1.bar(x + width / 2, self.sumsCur.values(), width, label=f"з/п {self.name}")

        ax1.grid(True, axis="y")
        ax1.set_title("Уровень зарплат по годам")
        ax1.set_xticks(np.arange(len(self.sums.keys())), self.sums.keys(), rotation=90)
        ax1.bar_label(remuneration1, fmt="")
        ax1.bar_label(remuneration2, fmt="")
        ax1.legend(prop={"size": 6})

        ax2.grid(True, axis="y")
        ax2.set_title("Количество вакансий по годам")
        x = np.arange(len(self.sums.keys()))
        ax2.set_xticks(x, self.sums.keys(), rotation=90)
        vacancyOne = ax2.bar(x - width / 2, self.sums.values(), width, label="Количество вакансий")
        vacancyTwo = ax2.bar(x + width / 2, self.sumsCur.values(), width, label=f"Количество вакансий\n{self.name}")
        ax2.bar_label(vacancyOne, fmt="")
        ax2.bar_label(vacancyTwo, fmt="")
        ax2.legend(prop={"size": 6})

        ax3.grid(True, axis="x")
        y = np.arange(len(list(self.ansSums.keys())))
        ax3.set_yticks(y, map(lambda s: s.replace(" ", "\n").replace("-", "\n"), self.ansSums.keys()))
        ax3.invert_yaxis()
        ax3.barh(y, self.ansSums.values())
        ax3.set_title("Уровень зарплат по городам")

        ax4.set_title("Доля вакансий по городам")
        other = 1 - sum(self.part.values())
        ax4.pie([other] + list(self.part.values()),
                labels=["Другие"] + list(self.part.keys()), startangle=0)

        fig.tight_layout(pad=0.4, w_pad=0.5, h_pad=1.0)
        plt.savefig("graph.png")
        plt.show()


fileName = input("Введите название файла: ")
profession = input("Введите название профессии: ")
report = Report(fileName, profession)
report.print_file()
report.generate_image()